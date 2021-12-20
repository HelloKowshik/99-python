from openpyxl import load_workbook

inv_file = load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']
products_per_suplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    unit_price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation for number of products per supplier
    if supplier_name in products_per_suplier:
        num_of_sup = products_per_suplier.get(supplier_name)
        products_per_suplier[supplier_name] = num_of_sup + 1
    else:
        products_per_suplier[supplier_name] = 1    

    # Calculation total value per supplier
    if supplier_name in total_value_per_supplier:
        current_price = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_price + (inventory * unit_price)
    else:
        total_value_per_supplier[supplier_name] = inventory * unit_price    

    # Calculation Products inventory less then 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Add value for total inventory price
    inventory_price.value = inventory * unit_price    

print(products_per_suplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save('new_inventory_file.xlsx')
